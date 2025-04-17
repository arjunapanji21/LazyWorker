import pandas as pd
import os
from openpyxl import load_workbook

class ExcelHandler:
    def __init__(self, filepath):
        self.filepath = filepath
        self.data = None
        self.sheet_name = None
        self.workbook = None
        self._column_cache = {}  # Add column cache
        self.chunk_size = 1000  # For batch processing
        
    def validate_file(self):
        """Validate Excel file existence and format"""
        if not self.filepath:
            raise ValueError("No Excel file selected")
        
        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")
            
        if not self.filepath.endswith(('.xlsx', '.xls')):
            raise ValueError("File must be an Excel file (.xlsx or .xls)")
    
    def get_sheet_names(self):
        """Get list of sheet names from Excel file"""
        try:
            self.workbook = load_workbook(self.filepath, read_only=True)
            return self.workbook.sheetnames
        except Exception as e:
            raise Exception(f"Error reading Excel sheets: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def get_data(self, sheet_name=None):
        """Optimized data loading with memory efficiency"""
        try:
            self.validate_file()
            
            # Use cached data if available
            if self.data is not None and self.sheet_name == sheet_name:
                return self.data
            
            sheets = self.get_sheet_names()
            if not sheets:
                raise ValueError("Excel file contains no sheets")
            
            self.sheet_name = sheet_name if sheet_name in sheets else sheets[0]
            
            # Read Excel with optimized settings
            self.data = pd.read_excel(
                self.filepath,
                sheet_name=self.sheet_name,
                dtype=str,  # Convert all columns to string
                engine='openpyxl',  # Use openpyxl engine
                na_filter=False  # Don't interpret anything as NaN
            )
            
            # Clean data
            self.data = self.data.fillna('')
            self.data = self.data.astype(str)
            
            if self.data.empty:
                raise ValueError(f"No data found in sheet: {self.sheet_name}")
            
            # Cache column names
            self._column_cache[self.sheet_name] = list(self.data.columns)
            
            return self.data
            
        except Exception as e:
            raise Exception(f"Error loading Excel data: {str(e)}")
    
    def get_column_names(self):
        """Get list of column names from current data"""
        if self.data is not None:
            return list(self.data.columns)
        return []
